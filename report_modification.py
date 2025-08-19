#!/usr/bin/env python3
import sys
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 4:
        print("Usage: remove_column_preserve_format.py <input.xlsx> <output.xlsx> <column_name>")
        sys.exit(1)

    input_file, output_file, column_to_remove = sys.argv[1:4]

    try:
        wb = load_workbook(input_file)
        ws = wb.active  # assumes first sheet

        # simple error case
        if not ws:
            print("Something went wrong with loading file.")
            return

        header_row = 19  # your actual header row
        headers = [cell.value for cell in ws[header_row]]

        if column_to_remove not in headers:
            print(f"Warning: column '{column_to_remove}' not found. File will be copied unchanged.")
            wb.save(output_file)
            return

        # Find column index (1-based)
        col_index = headers.index(column_to_remove) + 1

        # Delete that column
        ws.delete_cols(col_index)
        print(f"Removed column: {column_to_remove}")

        wb.save(output_file)
        print(f"New file written to: {output_file}")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
